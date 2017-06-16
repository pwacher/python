import openpyxl
import csv
import datetime

# get datetime
dateNow = datetime.datetime.now().strftime("%y%m%d - %H")

# Load in Excel Workbook & Sheet info and get ma number of rows
wb = openpyxl.load_workbook('dc7a.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
maxRows = sheet.max_row

# Loop through each cell
# a,b,c,d,e = 0
a = 0
b = 0
c = 0
d = 0
e = 0

# Ignore the header row, start at row 2 and loop through each row and count values into various buckets
for i in range(2, maxRows):
    num = sheet.cell(row=i, column=7).value
    if num >= 500:
    	a += 1
    elif num >= 200:
    	b += 1
    elif num >= 100:
    	c += 1
    elif num >= 50:
    	d += 1
    else:
    	e += 1

# Print results to terminal
print(dateNow)
print('>=500: ' + str(a))
print('>=200: ' + str(b))
print('>=100: ' + str(c))
print('>=50: ' + str(d))
print('<50: ' + str(e))

# Append values to a existing .csv file
# spreadsheet = csv.writer(open('dcSum.csv', 'a'), delimiter=',')
# spreadsheet.writerow([dateNow, a, b, c, d, e])
